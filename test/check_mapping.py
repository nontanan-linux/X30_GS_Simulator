import json
from openpyxl import load_workbook

wb83 = load_workbook('nestle_cat_waypoints_updated.xlsx', data_only=True)
ws83 = wb83.active
names83 = [str(row[1]).strip() for row in ws83.iter_rows(min_row=2, values_only=True) if row[1]]

with open('gs_cat/src/x30_udp_bridge/path/wet_zone.json', 'r') as f:
    json_data = json.load(f)

json_pts = [node['Node_info'] for node in json_data]

print(f'Total XLSX points: {len(names83)}')
print(f'Total JSON points: {len(json_pts)}')

# Just print them to see the alignment
print('XLSX INDEX | XLSX NAME | JSON INDEX | JSON NAME')
for i in range(min(50, len(names83))):
    # This might not align at all if there are Vias in between.
    # Let's find each mission point in the JSON sequentially.
    pass

# Better approach: find indices in JSON
current_json_idx = 0
for i, name in enumerate(names83):
    found = False
    name_clean = name.replace(' ', '_').lower()
    for j in range(current_json_idx, len(json_pts)):
        json_name_clean = json_pts[j].replace(' ', '_').lower()
        # Handle cases like loto01 vs loto1
        if json_name_clean == name_clean or json_name_clean == name_clean.replace('01', '1').replace('02', '2').replace('03', '3').replace('04', '4').replace('05', '5').replace('06', '6').replace('07', '7').replace('08', '8').replace('09', '9'):
            print(f'{i:3} | {name:30} | {j:4} | {json_pts[j]}')
            current_json_idx = j + 1
            found = True
            break
        # Special case for Home
        if name.lower() == 'home' and (json_name_clean == 'home' or json_name_clean == 'test' or json_name_clean == ''):
            if i == 0: # Usually the first point
                print(f'{i:3} | {name:30} | {j:4} | {json_pts[j]}')
                current_json_idx = j + 1
                found = True
                break
    if not found:
        print(f'{i:3} | {name:30} | NOT FOUND')
